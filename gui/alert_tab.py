"""
Alert/Review tab UI - displays alerts in a treeview with bulk actions.

This tab shows comparison results that meet similarity/profit thresholds
and allows users to manage the reselling workflow.
"""

import csv
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import webbrowser
import logging
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional

from gui.alert_manager import AlertManager
from gui.alert_states import (
    AlertState,
    AlertBulkActions,
    get_state_color,
    get_state_display_name
)


class AlertTab(ttk.Frame):
    """Alert/Review tab for managing items through reselling workflow."""

    def __init__(self, parent, settings_manager=None):
        """
        Initialize Alert tab.

        Args:
            parent: Parent widget
            settings_manager: Optional settings manager instance
        """
        super().__init__(parent)
        self.selected_alert_ids = []
        self.settings_manager = settings_manager

        # Load saved settings or use defaults
        if settings_manager:
            alert_settings = settings_manager.get_alert_settings()
            min_sim = alert_settings.get('filter_min_similarity', 70.0)
            min_profit = alert_settings.get('filter_min_profit', 20.0)
            self.notifications_enabled = alert_settings.get('notifications_enabled', False)
            self.notify_min_sim = alert_settings.get('notify_min_similarity', 80.0)
            self.notify_min_profit = alert_settings.get('notify_min_profit', 30.0)
        else:
            min_sim = 70.0
            min_profit = 20.0
            self.notifications_enabled = False
            self.notify_min_sim = 80.0
            self.notify_min_profit = 30.0

        # Initialize alert manager with notification settings
        self.alert_manager = AlertManager(
            notifications_enabled=self.notifications_enabled,
            notify_min_similarity=self.notify_min_sim,
            notify_min_profit=self.notify_min_profit
        )

        # Threshold variables
        self.min_similarity_var = tk.DoubleVar(value=min_sim)
        self.min_profit_var = tk.DoubleVar(value=min_profit)

        # UI state
        self.state_filter_var = tk.StringVar(value="all")

        # Sorting state
        self.sort_column = None
        self.sort_reverse = False
        self.state_sort_index = 0  # For cycling through states

        self._build_ui()
        self._load_alerts()

    def _build_ui(self):
        """Build the UI components."""
        # Top controls frame - everything on the left
        controls_frame = ttk.Frame(self)
        controls_frame.pack(fill=tk.X, padx=5, pady=5)

        # All filters on the left side
        filters_frame = ttk.LabelFrame(controls_frame, text="Filters:", padding=5)
        filters_frame.pack(side=tk.LEFT, padx=(0, 5))

        # State filter
        ttk.Label(filters_frame, text="State:").pack(side=tk.LEFT, padx=5)
        state_combo = ttk.Combobox(
            filters_frame,
            textvariable=self.state_filter_var,
            values=["all", "pending", "yay", "nay", "purchased", "received", "posted", "sold"],
            state="readonly",
            width=12
        )
        state_combo.pack(side=tk.LEFT, padx=5)
        state_combo.bind("<<ComboboxSelected>>", lambda e: self._load_alerts())

        # Separator
        ttk.Separator(filters_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        # Similarity filter
        ttk.Label(filters_frame, text="Min Similarity %:").pack(side=tk.LEFT, padx=5)
        similarity_spinbox = ttk.Spinbox(
            filters_frame,
            from_=0,
            to=100,
            textvariable=self.min_similarity_var,
            width=8
        )
        similarity_spinbox.pack(side=tk.LEFT, padx=5)
        self.min_similarity_var.trace_add("write", lambda *args: self._on_filter_change())

        # Profit filter
        ttk.Label(filters_frame, text="Min Profit %:").pack(side=tk.LEFT, padx=5)
        profit_spinbox = ttk.Spinbox(
            filters_frame,
            from_=-100,
            to=1000,
            textvariable=self.min_profit_var,
            width=8
        )
        profit_spinbox.pack(side=tk.LEFT, padx=5)
        self.min_profit_var.trace_add("write", lambda *args: self._on_filter_change())

        # Notification settings (on right side of filters)
        notify_frame = ttk.LabelFrame(controls_frame, text="Notifications:", padding=5)
        notify_frame.pack(side=tk.LEFT, padx=(10, 0))

        self.notify_enabled_var = tk.BooleanVar(value=self.notifications_enabled)
        notify_check = ttk.Checkbutton(
            notify_frame,
            text="Enable",
            variable=self.notify_enabled_var,
            command=self._on_notification_toggle
        )
        notify_check.pack(side=tk.LEFT, padx=5)

        ttk.Label(notify_frame, text="Min Sim:").pack(side=tk.LEFT, padx=(10, 2))
        self.notify_sim_var = tk.DoubleVar(value=self.notify_min_sim)
        notify_sim_spin = ttk.Spinbox(
            notify_frame,
            from_=0,
            to=100,
            textvariable=self.notify_sim_var,
            width=6,
            command=self._on_notification_settings_change
        )
        notify_sim_spin.pack(side=tk.LEFT, padx=2)
        ttk.Label(notify_frame, text="%").pack(side=tk.LEFT)

        ttk.Label(notify_frame, text="Min Profit:").pack(side=tk.LEFT, padx=(10, 2))
        self.notify_profit_var = tk.DoubleVar(value=self.notify_min_profit)
        notify_profit_spin = ttk.Spinbox(
            notify_frame,
            from_=-100,
            to=1000,
            textvariable=self.notify_profit_var,
            width=6,
            command=self._on_notification_settings_change
        )
        notify_profit_spin.pack(side=tk.LEFT, padx=2)
        ttk.Label(notify_frame, text="%").pack(side=tk.LEFT)

        # Bulk actions frame
        actions_frame = ttk.Frame(self)
        actions_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(actions_frame, text="Bulk Actions:").pack(side=tk.LEFT, padx=5)

        # Action buttons
        ttk.Button(actions_frame, text="Mark Yay", command=lambda: self._bulk_action("yay")).pack(side=tk.LEFT, padx=2)
        ttk.Button(actions_frame, text="Mark Nay", command=lambda: self._bulk_action("nay")).pack(side=tk.LEFT, padx=2)
        ttk.Separator(actions_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(actions_frame, text="Purchase", command=lambda: self._bulk_action("purchased")).pack(side=tk.LEFT, padx=2)
        ttk.Button(actions_frame, text="Received", command=lambda: self._bulk_action("received")).pack(side=tk.LEFT, padx=2)
        ttk.Button(actions_frame, text="Posted", command=lambda: self._bulk_action("posted")).pack(side=tk.LEFT, padx=2)
        ttk.Button(actions_frame, text="Sold", command=lambda: self._bulk_action("sold")).pack(side=tk.LEFT, padx=2)

        ttk.Separator(actions_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        ttk.Button(actions_frame, text="Delete Selected", command=self._delete_selected).pack(side=tk.LEFT, padx=2)

        ttk.Separator(actions_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        ttk.Button(actions_frame, text="Export to Spreadsheet", command=self._export_alerts).pack(side=tk.LEFT, padx=2)

        # Treeview frame
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Scrollbars
        tree_scroll_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        tree_scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        # Treeview columns
        columns = (
            "state",
            "similarity",
            "profit",
            "store_title",
            "ebay_title",
            "store_price",
            "ebay_price",
            "shipping",
            "sold_date"
        )

        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="tree headings",
            selectmode="extended",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )
        self.tree.pack(fill=tk.BOTH, expand=True)

        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)

        # Configure columns
        self.tree.column("#0", width=50, minwidth=50)  # ID column
        self.tree.column("state", width=100, minwidth=80)
        self.tree.column("similarity", width=80, minwidth=60)
        self.tree.column("profit", width=80, minwidth=60)
        self.tree.column("store_title", width=250, minwidth=150)
        self.tree.column("ebay_title", width=250, minwidth=150)
        self.tree.column("store_price", width=100, minwidth=80)
        self.tree.column("ebay_price", width=80, minwidth=60)
        self.tree.column("shipping", width=80, minwidth=60)
        self.tree.column("sold_date", width=100, minwidth=80)

        # Headings with sort commands
        self.tree.heading("#0", text="ID", command=lambda: self._sort_column("#0", False))
        self.tree.heading("state", text="State", command=lambda: self._sort_by_state())
        self.tree.heading("similarity", text="Similarity %", command=lambda: self._sort_column("similarity", False))
        self.tree.heading("profit", text="Profit %", command=lambda: self._sort_column("profit", False))
        self.tree.heading("store_title", text="Store Title", command=lambda: self._sort_column("store_title", False))
        self.tree.heading("ebay_title", text="eBay Title", command=lambda: self._sort_column("ebay_title", False))
        self.tree.heading("store_price", text="Store Price", command=lambda: self._sort_column("store_price", False))
        self.tree.heading("ebay_price", text="eBay Price", command=lambda: self._sort_column("ebay_price", False))
        self.tree.heading("shipping", text="Shipping", command=lambda: self._sort_column("shipping", False))
        self.tree.heading("sold_date", text="Sold Date", command=lambda: self._sort_column("sold_date", False))

        # Bind events
        self.tree.bind("<Double-Button-1>", self._on_double_click)
        self.tree.bind("<Button-3>", self._on_right_click)

        # Status bar
        status_frame = ttk.Frame(self)
        status_frame.pack(fill=tk.X, padx=5, pady=5)

        self.status_label = ttk.Label(status_frame, text="No alerts loaded")
        self.status_label.pack(side=tk.LEFT)

    def _on_filter_change(self):
        """Handle filter value changes - save settings and reload alerts."""
        # Save settings if settings manager available
        if self.settings_manager:
            try:
                self.settings_manager.save_alert_settings(
                    filter_min_similarity=self.min_similarity_var.get(),
                    filter_min_profit=self.min_profit_var.get()
                )
            except Exception as e:
                logging.warning(f"Failed to save alert filter settings: {e}")

        # Reload alerts with new filters
        self._load_alerts()

    def _on_notification_toggle(self):
        """Handle notification enable/disable toggle."""
        enabled = self.notify_enabled_var.get()

        # Update alert manager notification state
        self.alert_manager.notifier.enabled = enabled
        self.alert_manager.notification_filter.enabled = enabled

        # Save to settings
        if self.settings_manager:
            try:
                settings = self.settings_manager.get_alert_settings()
                settings['notifications_enabled'] = enabled
                self.settings_manager.save_alert_settings(**settings)
                logging.info(f"Notifications {'enabled' if enabled else 'disabled'}")
            except Exception as e:
                logging.warning(f"Failed to save notification settings: {e}")

    def _on_notification_settings_change(self):
        """Handle notification threshold changes."""
        notify_min_sim = self.notify_sim_var.get()
        notify_min_profit = self.notify_profit_var.get()

        # Update alert manager notification thresholds
        self.alert_manager.notification_filter.min_similarity = notify_min_sim
        self.alert_manager.notification_filter.min_profit = notify_min_profit

        # Save to settings
        if self.settings_manager:
            try:
                settings = self.settings_manager.get_alert_settings()
                settings['notify_min_similarity'] = notify_min_sim
                settings['notify_min_profit'] = notify_min_profit
                self.settings_manager.save_alert_settings(**settings)
            except Exception as e:
                logging.warning(f"Failed to save notification threshold settings: {e}")

    def _ask_yes_no_at_cursor(self, title: str, message: str) -> bool:
        """Show yes/no dialog at cursor position."""
        # Create a custom dialog
        dialog = tk.Toplevel(self)
        dialog.title(title)
        dialog.resizable(False, False)
        dialog.transient(self)

        # Get mouse position
        x = self.winfo_pointerx()
        y = self.winfo_pointery()

        # Position dialog at cursor
        dialog.geometry(f"+{x}+{y}")

        # Result variable
        result = [False]

        # Message label
        tk.Label(dialog, text=message, padx=20, pady=20).pack()

        # Button frame
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=(0, 10))

        def on_yes():
            result[0] = True
            dialog.destroy()

        def on_no():
            result[0] = False
            dialog.destroy()

        # Yes/No buttons
        ttk.Button(button_frame, text="Yes", command=on_yes, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="No", command=on_no, width=10).pack(side=tk.LEFT, padx=5)

        # Make dialog modal
        dialog.grab_set()
        dialog.focus_set()

        # Bind Enter/Escape keys
        dialog.bind('<Return>', lambda e: on_yes())
        dialog.bind('<Escape>', lambda e: on_no())

        # Wait for dialog to close
        dialog.wait_window()

        return result[0]

    def _sort_by_state(self):
        """Sort by state column - cycles through state order."""
        # State order for cycling
        state_order = [
            AlertState.PENDING,
            AlertState.YAY,
            AlertState.NAY,
            AlertState.PURCHASED,
            AlertState.RECEIVED,
            AlertState.POSTED,
            AlertState.SOLD
        ]

        # Get current filter
        state_filter_str = self.state_filter_var.get()
        if state_filter_str != "all":
            try:
                current_state = AlertState(state_filter_str)
                # If we're already filtered by this state, cycle to next
                current_idx = state_order.index(current_state)
                next_idx = (current_idx + 1) % len(state_order)
                self.state_filter_var.set(state_order[next_idx].value)
            except ValueError:
                # Invalid state, start from beginning
                self.state_filter_var.set(state_order[0].value)
        else:
            # Start filtering from first state
            self.state_filter_var.set(state_order[0].value)

        # Reload with new filter
        self._load_alerts()

    def _sort_column(self, col, is_numeric):
        """Sort treeview by column."""
        # Toggle sort direction if clicking same column
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False

        # Get all items with their data
        items = []
        for child in self.tree.get_children():
            if col == "#0":
                # Sort by ID (text column)
                value = self.tree.item(child, "text")
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    value = 0
            else:
                # Get column index
                col_index = {
                    "state": 0,
                    "similarity": 1,
                    "profit": 2,
                    "store_title": 3,
                    "ebay_title": 4,
                    "store_price": 5,
                    "ebay_price": 6,
                    "shipping": 7,
                    "sold_date": 8
                }.get(col, 0)

                values = self.tree.item(child, "values")
                value = values[col_index] if col_index < len(values) else ""

                # Extract numeric values from percentage strings
                if col in ["similarity", "profit"]:
                    try:
                        value = float(value.replace("%", "").strip()) if value != "-" else -999999
                    except (ValueError, AttributeError):
                        value = -999999
                elif col in ["store_price", "ebay_price", "shipping"]:
                    # Extract numeric value from price strings
                    try:
                        value = float(value.replace("¥", "").replace("$", "").replace(",", "").strip())
                    except (ValueError, AttributeError):
                        value = 0

            items.append((value, child))

        # Sort items
        items.sort(reverse=self.sort_reverse)

        # Rearrange items in tree
        for index, (value, child) in enumerate(items):
            self.tree.move(child, "", index)

        # Update column heading to show sort direction
        self._update_column_headings()

    def _update_column_headings(self):
        """Update column headings to show sort indicator."""
        # Column display names and commands
        column_config = {
            "#0": ("ID", lambda: self._sort_column("#0", False)),
            "state": ("State", lambda: self._sort_by_state()),
            "similarity": ("Similarity %", lambda: self._sort_column("similarity", False)),
            "profit": ("Profit %", lambda: self._sort_column("profit", False)),
            "store_title": ("Store Title", lambda: self._sort_column("store_title", False)),
            "ebay_title": ("eBay Title", lambda: self._sort_column("ebay_title", False)),
            "store_price": ("Store Price", lambda: self._sort_column("store_price", False)),
            "ebay_price": ("eBay Price", lambda: self._sort_column("ebay_price", False)),
            "shipping": ("Shipping", lambda: self._sort_column("shipping", False)),
            "sold_date": ("Sold Date", lambda: self._sort_column("sold_date", False))
        }

        # Update all headings
        for col, (heading, command) in column_config.items():
            if col == self.sort_column:
                # Add sort indicator
                indicator = " ▼" if self.sort_reverse else " ▲"
                self.tree.heading(col, text=heading + indicator, command=command)
            else:
                self.tree.heading(col, text=heading, command=command)

    def _load_alerts(self):
        """Load alerts from storage and populate treeview with filters applied."""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Get state filter
        state_filter_str = self.state_filter_var.get()
        state_filter = None
        if state_filter_str != "all":
            try:
                state_filter = AlertState(state_filter_str)
            except ValueError:
                pass

        # Load alerts
        if state_filter:
            alerts = self.alert_manager.get_alerts_by_state(state_filter)
        else:
            alerts = self.alert_manager.get_all_alerts()

        # Apply similarity and profit filters (for display only)
        try:
            min_similarity = self.min_similarity_var.get()
            min_profit = self.min_profit_var.get()

            filtered_alerts = [
                alert for alert in alerts
                if alert.get('similarity', 0) >= min_similarity and
                   alert.get('profit_margin', 0) >= min_profit
            ]
        except (ValueError, TypeError, tk.TclError) as e:
            # If filter values are invalid, show all
            logging.warning(f"Invalid filter values, showing all alerts: {e}")
            filtered_alerts = alerts

        # Sort by ID (most recent first)
        filtered_alerts.sort(key=lambda x: x.get('alert_id', 0), reverse=True)

        # Populate treeview
        for alert in filtered_alerts:
            self._add_alert_to_tree(alert)

        # Update status
        if len(filtered_alerts) < len(alerts):
            self.status_label.config(text=f"Showing {len(filtered_alerts)} of {len(alerts)} alerts (filtered)")
        else:
            self.status_label.config(text=f"Loaded {len(alerts)} alerts")

        # Reapply sort if one was active
        if self.sort_column:
            self._sort_column(self.sort_column, False)
        else:
            self._update_column_headings()

    def _add_alert_to_tree(self, alert: Dict):
        """Add a single alert to the treeview."""
        alert_id = alert.get('alert_id', 0)
        state = alert.get('state', AlertState.PENDING)

        # Format values
        similarity = alert.get('similarity', 0)
        similarity_str = f"{similarity:.1f}%" if similarity > 0 else "-"

        profit = alert.get('profit_margin', 0)
        profit_str = f"{profit:.1f}%" if profit != 0 else "-"

        values = (
            get_state_display_name(state),
            similarity_str,
            profit_str,
            alert.get('store_title', 'N/A'),
            alert.get('ebay_title', 'N/A'),
            alert.get('store_price', '¥0'),
            alert.get('ebay_price', '$0'),
            alert.get('shipping', '$0'),
            alert.get('sold_date', '')
        )

        # Insert item
        item_id = self.tree.insert("", "end", text=str(alert_id), values=values, tags=(state.value,))

        # Color-code by state
        color = get_state_color(state)
        self.tree.tag_configure(state.value, background=color)

    def _bulk_action(self, action: str):
        """
        Perform bulk action on selected items.

        Args:
            action: Action name (yay, nay, purchased, received, etc.)
        """
        selected_items = self.tree.selection()
        if not selected_items:
            return

        # Get alert IDs
        alert_ids = [int(self.tree.item(item, "text")) for item in selected_items]

        # Map action to state
        state_map = {
            "yay": AlertState.YAY,
            "nay": AlertState.NAY,
            "purchased": AlertState.PURCHASED,
            "received": AlertState.RECEIVED,
            "posted": AlertState.POSTED,
            "sold": AlertState.SOLD
        }

        new_state = state_map.get(action)
        if not new_state:
            return

        # Special handling for "Purchase" state - open all store URLs in browser tabs
        if new_state == AlertState.PURCHASED:
            alerts_data = self.alert_manager.get_alerts_by_ids(alert_ids)
            opened_count = 0
            for alert in alerts_data:
                url = alert.get('store_link', '')
                if url:
                    webbrowser.open(url)
                    opened_count += 1
            if opened_count > 0:
                print(f"[PURCHASE] Opened {opened_count} store URLs in browser")

        # Perform bulk update
        success_count = self.alert_manager.bulk_update_state(alert_ids, new_state)

        # Special handling for "Posted" state - create eBay draft listings
        if new_state == AlertState.POSTED:
            self._create_ebay_listings(alert_ids)

        # Reload
        self._load_alerts()

    def _delete_selected(self):
        """Delete selected alerts."""
        selected_items = self.tree.selection()
        if not selected_items:
            return

        alert_ids = [int(self.tree.item(item, "text")) for item in selected_items]

        if not self._ask_yes_no_at_cursor("Confirm Delete", f"Delete {len(alert_ids)} alerts? This cannot be undone."):
            return

        self.alert_manager.delete_alerts(alert_ids)
        self._load_alerts()

    def _on_double_click(self, event):
        """Handle double-click on alert item - open links based on clicked column."""
        item = self.tree.identify_row(event.y)
        if not item:
            return

        # Identify which column was clicked
        column = self.tree.identify_column(event.x)

        alert_id = int(self.tree.item(item, "text"))
        alert = self.alert_manager.storage.get_alert_by_id(alert_id)

        if not alert:
            return

        # Open appropriate link based on column
        if column == "#4":  # store_title column (0-indexed: #0=ID, #1=state, #2=similarity, #3=profit, #4=store_title)
            link = alert.get('store_link', '')
            if link:
                webbrowser.open(link)
        elif column == "#5":  # ebay_title column
            link = alert.get('ebay_link', '')
            if link:
                webbrowser.open(link)

    def _on_right_click(self, event):
        """Handle right-click - show context menu."""
        item = self.tree.identify_row(event.y)
        if not item:
            return

        # Create context menu
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Open Store Link", command=lambda: self._open_store_link(item))
        menu.add_command(label="Open eBay Link", command=lambda: self._open_ebay_link(item))
        menu.add_separator()
        menu.add_command(label="Mark as Yay", command=lambda: self._quick_action(item, AlertState.YAY))
        menu.add_command(label="Mark as Nay", command=lambda: self._quick_action(item, AlertState.NAY))
        menu.add_separator()
        menu.add_command(label="Delete", command=lambda: self._delete_item(item))

        menu.post(event.x_root, event.y_root)

    def _open_store_link(self, item):
        """Open store link for item."""
        alert_id = int(self.tree.item(item, "text"))
        alert = self.alert_manager.storage.get_alert_by_id(alert_id)
        if alert and alert.get('store_link'):
            webbrowser.open(alert['store_link'])

    def _open_ebay_link(self, item):
        """Open eBay link for item."""
        alert_id = int(self.tree.item(item, "text"))
        alert = self.alert_manager.storage.get_alert_by_id(alert_id)
        if alert and alert.get('ebay_link'):
            webbrowser.open(alert['ebay_link'])

    def _quick_action(self, item, state: AlertState):
        """Quick action from context menu."""
        alert_id = int(self.tree.item(item, "text"))
        self.alert_manager.update_alert_state(alert_id, state)
        self._load_alerts()

    def _delete_item(self, item):
        """Delete selected items (or single item if only one selected)."""
        selected_items = self.tree.selection()

        # If multiple items selected, delete all of them
        if len(selected_items) > 1:
            alert_ids = [int(self.tree.item(i, "text")) for i in selected_items]
            if self._ask_yes_no_at_cursor("Confirm Delete", f"Delete {len(alert_ids)} alerts? This cannot be undone."):
                self.alert_manager.delete_alerts(alert_ids)
                self._load_alerts()
        else:
            # Single item delete
            alert_id = int(self.tree.item(item, "text"))
            if self._ask_yes_no_at_cursor("Confirm Delete", f"Delete alert #{alert_id}?"):
                self.alert_manager.delete_alerts([alert_id])
                self._load_alerts()

    def _create_ebay_listings(self, alert_ids: List[int]):
        """
        Create eBay draft listings for selected alerts.

        Args:
            alert_ids: List of alert IDs to create listings for
        """
        try:
            from ebay_listing_creator import create_listing_from_alert

            alerts_data = self.alert_manager.get_alerts_by_ids(alert_ids)
            created_count = 0
            failed_count = 0

            for alert in alerts_data:
                try:
                    # Convert price from display format (e.g., "$29.99") to float
                    ebay_price_str = alert.get('ebay_price', '$0')
                    ebay_price = float(ebay_price_str.replace('$', '').replace(',', ''))

                    # Prepare alert data for listing creator
                    listing_data = {
                        'id': str(alert.get('alert_id', alert.get('id', 'unknown'))),
                        'store_title': alert.get('store_title', 'Untitled'),
                        'store_title_en': alert.get('store_title_en', alert.get('store_title', 'Untitled')),
                        'ebay_price': ebay_price,
                        'store_link': alert.get('store_link', ''),
                        'store_images': alert.get('store_images', [])
                    }

                    result = create_listing_from_alert(listing_data)
                    if result:
                        created_count += 1
                        print(f"[EBAY LISTING] ✓ Created listing for alert #{listing_data['id']}")
                    else:
                        failed_count += 1
                        print(f"[EBAY LISTING] ✗ Failed for alert #{listing_data['id']}")

                except Exception as e:
                    failed_count += 1
                    print(f"[EBAY LISTING] Error creating listing: {e}")
                    import traceback
                    traceback.print_exc()

            # Print summary to console
            if created_count > 0 or failed_count > 0:
                print(f"[EBAY LISTING] Created: {created_count}, Failed: {failed_count}")

        except ImportError:
            print("[EBAY LISTING] Error: eBay listing creator module not found")
        except Exception as e:
            print(f"[EBAY LISTING] Error: Failed to create eBay listings: {e}")

    def _export_alerts(self):
        """Export currently visible alerts to CSV or Excel."""
        # Get currently visible items from treeview
        items = self.tree.get_children()

        if not items:
            messagebox.showinfo("Export Alerts", "No alerts to export.")
            return

        # Ask user for file location and format
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ],
            initialfile=f"alerts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )

        if not file_path:
            return  # User cancelled

        try:
            # Gather all alert data from visible items
            export_data = []
            for item_id in items:
                values = self.tree.item(item_id, 'values')
                alert_id = self.tree.item(item_id, 'text')

                # Get full alert data from manager
                alert = self.alert_manager.storage.get_alert_by_id(int(alert_id))

                if alert:
                    export_data.append({
                        'Alert ID': alert_id,
                        'State': values[0] if len(values) > 0 else '',
                        'Similarity %': values[1] if len(values) > 1 else '',
                        'Profit %': values[2] if len(values) > 2 else '',
                        'Store Title': values[3] if len(values) > 3 else '',
                        'eBay Title': values[4] if len(values) > 4 else '',
                        'Store Price': values[5] if len(values) > 5 else '',
                        'eBay Price': values[6] if len(values) > 6 else '',
                        'Shipping': values[7] if len(values) > 7 else '',
                        'Sold Date': values[8] if len(values) > 8 else '',
                        'Store Link': alert.get('store_link', ''),
                        'eBay Link': alert.get('ebay_link', ''),
                        'Created At': alert.get('created_at', ''),
                        'Updated At': alert.get('updated_at', '')
                    })

            # Export based on file extension
            file_ext = Path(file_path).suffix.lower()

            if file_ext == '.xlsx':
                self._export_to_excel(export_data, file_path)
            else:
                self._export_to_csv(export_data, file_path)

            messagebox.showinfo(
                "Export Complete",
                f"Successfully exported {len(export_data)} alerts to:\n{file_path}"
            )

        except Exception as e:
            logging.error(f"Failed to export alerts: {e}")
            messagebox.showerror("Export Error", f"Failed to export alerts:\n{e}")

    def _export_to_csv(self, data: List[Dict], file_path: str):
        """Export data to CSV file."""
        if not data:
            return

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

    def _export_to_excel(self, data: List[Dict], file_path: str):
        """Export data to Excel file using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Alerts"

            if not data:
                wb.save(file_path)
                return

            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

            # Auto-size columns
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(header))
                for row_idx in range(2, len(data) + 2):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            wb.save(file_path)

        except ImportError:
            # Fall back to CSV if openpyxl not installed
            logging.warning("openpyxl not installed, falling back to CSV export")
            csv_path = str(Path(file_path).with_suffix('.csv'))
            self._export_to_csv(data, csv_path)
            raise ImportError(f"Excel export requires openpyxl. Exported to CSV instead: {csv_path}")

    def add_filtered_alerts(self, comparison_results: List[Dict]):
        """
        Add pre-filtered comparison results to alerts.

        Called from eBay Search tab with results already filtered by thresholds.

        Args:
            comparison_results: List of comparison result dictionaries (already filtered)
        """
        # Process results without applying additional thresholds (already filtered)
        created_alerts = self.alert_manager.process_comparison_results(
            comparison_results,
            min_similarity=0,  # No additional filtering
            min_profit=-999999  # No additional filtering
        )

        # Reload alerts to show new ones
        self._load_alerts()
